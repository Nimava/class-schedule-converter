import pandas as pd
import os
import re
from math import ceil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
import hashlib
import tkinter as tk
from tkinter import filedialog, messagebox
import sys
import tempfile

def show_welcome_message():
    """Show welcome message before file selection"""
    root = tk.Tk()
    root.withdraw()
    
    welcome_text = """برنامه تبدیل خروجی آموزشیار به اکسل کاشی کلاسها

با توجه به امکان تغییر در خروجی آموزشیار در بروزرسانی، لطفا از آخرین نسخه برنامه استفاده نمایید.

نسخه 1.3 - بهمن 1404 - نیما وزیری"""
    
    messagebox.showinfo("خوش آمدید", welcome_text)

def select_input_file():
    """Open file dialog to select input CSV file"""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.askopenfilename(
        title="لطفا فایل CSV را انتخاب کنید",
        filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
    )
    
    return file_path

def select_output_file():
    """Open file dialog to select output Excel file location"""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.asksaveasfilename(
        title="ذخیره فایل اکسل نهایی",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    return file_path

def phase1_extract_data(input_file, temp_output_file):
    """Phase 1: Extract important data from CSV and save to Excel"""
    print("📖 در حال خواندن فایل CSV ...")
    
    try:
        # ==== خواندن فایل ورودی ====
        df = pd.read_csv(input_file, encoding='utf-8-sig')
        print(f"✅ فایل خوانده شد. تعداد ردیف‌ها: {len(df)}")
        
        # ==== انتخاب ستون‌ها بر اساس شماره ====
        selected_columns = {
            'نام درس': 2,           # C
            'کد ارائه درس': 0,      # A
            'واحد نظری': 11,        # L
            'واحد عملی': 12,        # M
            'مکان': 22,             # W
            'گروه آموزشی': 43,      # AR
            'مقطع': 53,             # BB
            'تعداد ثبت نامی': 57,   # BF
            'نیم‌سال': 59,          # BH
            'نام استاد': 68,        # BQ
            'رشته': 70,             # BS
            'روز': 72,              # BU
            'ساعت شروع': 73,        # BV
            'ساعت پایان': 74,       # BW
            'تقويم كلاس درس': 71   # BT - اضافه شد
        }
        
        # ==== استخراج فقط ستون‌های مورد نیاز ====
        df_selected = df.iloc[:, list(selected_columns.values())].copy()
        df_selected.columns = list(selected_columns.keys())
        
        # ==== پاکسازی و نرمال‌سازی ====
        def normalize_text(s):
            return (
                str(s)
                .replace('\u200c', '')   # حذف نیم‌فاصله
                .replace('ي', 'ی')       # ی عربی → فارسی
                .replace('ك', 'ک')       # ک عربی → فارسی
                .replace('‌', '')        # حذف ZWNJ اضافی
                .strip()
            )
        
        df_selected = df_selected.fillna("").astype(str)
        
        # NEW: استخراج اطلاعات از ستون تقويم كلاس درس اگر ستون‌های روز و ساعت خالی باشند
        def extract_from_calendar(calendar_text):
            """استخراج روز، ساعت شروع و پایان از متن تقويم كلاس درس"""
            if not calendar_text or calendar_text.strip() == "":
                return "", "", ""
            
            text = str(calendar_text).strip()
            # نرمال‌سازی متن قبل از پردازش
            text = normalize_text(text)
            
            # جستجوی روز با الگوی دقیق
            day = ""
            
            # استفاده از regex برای شناسایی تمام اشکال ممکن
            patterns = [
                (r'^شنبه', 'شنبه'),
                (r'^یکشنبه', 'یکشنبه'),
                (r'^دوشنبه', 'دوشنبه'),
                (r'^سه[‌_\s]*شنبه', 'سه‌شنبه'),  # سه‌شنبه، سه_شنبه، سه شنبه، سهشنبه
                (r'^چهار[‌_\s]*شنبه', 'چهارشنبه'),  # چهارشنبه، چهار_شنبه، چهار شنبه
                (r'^پنج[‌_\s]*شنبه', 'پنج‌شنبه'),  # پنج‌شنبه، پنج_شنبه، پنج شنبه، پنجشنبه
                (r'^جمعه', 'جمعه')
            ]
            
            for pattern, day_name in patterns:
                if re.match(pattern, text, re.UNICODE):
                    day = day_name
                    break
            
            # جستجوی ساعت‌ها با الگوی "ساعت تا ساعت"
            time_pattern = r'(\d{1,2}[:\.]\d{2})\s*تا\s*(\d{1,2}[:\.]\d{2})'
            time_match = re.search(time_pattern, text)
            
            start_time = ""
            end_time = ""
            
            if time_match:
                start_time = time_match.group(1).replace('.', ':')
                end_time = time_match.group(2).replace('.', ':')
            
            return day, start_time, end_time
        
        # پردازش هر ردیف
        for idx, row in df_selected.iterrows():
            # اگر روز یا ساعت خالی باشد، از ستون تقويم كلاس درس استخراج کن
            if (row['روز'].strip() == "" or 
                row['ساعت شروع'].strip() == "" or 
                row['ساعت پایان'].strip() == ""):
                
                calendar_text = row['تقويم كلاس درس']
                day_from_cal, start_from_cal, end_from_cal = extract_from_calendar(calendar_text)
                
                if row['روز'].strip() == "" and day_from_cal:
                    df_selected.at[idx, 'روز'] = day_from_cal
                
                if row['ساعت شروع'].strip() == "" and start_from_cal:
                    df_selected.at[idx, 'ساعت شروع'] = start_from_cal
                
                if row['ساعت پایان'].strip() == "" and end_from_cal:
                    df_selected.at[idx, 'ساعت پایان'] = end_from_cal
        
        # نرمال‌سازی روزها (همانند قبل)
        df_selected['روز'] = df_selected['روز'].apply(normalize_text)
        
        # ==== نگاشت دقیق اسامی روزها ====
        day_map = {
            'شنبه': 'شنبه',
            'یکشنبه': 'یکشنبه',
            'يکشنبه': 'یکشنبه',
            'يكشنبه': 'یکشنبه',
            'یكشنبه': 'یکشنبه',
            'دوشنبه': 'دوشنبه',
            'سه شنبه': 'سه‌شنبه',
            'سه‌شنبه': 'سه‌شنبه',
            'سهشنبه': 'سه‌شنبه',  # اضافه شد
            'چهارشنبه': 'چهارشنبه',
            'چهار شنبه': 'چهارشنبه',
            'پنجشنبه': 'پنج‌شنبه',
            'پنج شنبه': 'پنج‌شنبه',
            'پنج‌شنبه': 'پنج‌شنبه',
            'پنچشنبه': 'پنج‌شنبه',      # حالت اشتباه تایپی احتمالی
            'پنچ شنبه': 'پنج‌شنبه',
            'جمعه': 'جمعه'
        }
        
        # 🔹 نگاشت با تطبیق دقیق (نه جستجوی درون رشته)
        df_selected['روز'] = df_selected['روز'].apply(
            lambda x: day_map[x] if x in day_map else x
        )
        
        # ==== لیست روزهای معتبر ====
        days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنج‌شنبه', 'جمعه']
        
        # ==== تقسیم داده‌ها به شیت‌های مجزا و مرتب‌سازی ====
        sheets = {}
        for day in days:
            subset = df_selected[df_selected['روز'] == day].copy()
            if not subset.empty:
                # مرتب‌سازی بر اساس ساعت شروع
                subset['ساعت شروع مرتب'] = subset['ساعت شروع'].str.extract(r'(\d+)').astype(float)
                subset = subset.sort_values(by='ساعت شروع مرتب', ascending=True).drop(columns=['ساعت شروع مرتب'])
                sheets[day] = subset
        
        # ==== داده‌های با روز نامشخص ====
        unknown = df_selected[~df_selected['روز'].isin(days)]
        if not unknown.empty:
            sheets['نامشخص'] = unknown
        
        # ==== ذخیره در فایل اکسل ====
        with pd.ExcelWriter(temp_output_file, engine='openpyxl') as writer:
            for day, subset in sheets.items():
                subset.to_excel(writer, sheet_name=day[:30], index=False)
        
        print("✅ فایل اکسل موقت ساخته شد:", temp_output_file)
        print("📅 روزهای شناسایی‌شده:", list(sheets.keys()))
        return True
        
    except Exception as e:
        print(f"❌ خطا در فاز اول: {e}")
        return False
        
def phase2_create_schedule(temp_file, final_output_file):
    """Phase 2: Create class schedule tables from the temporary Excel file"""
    
    # Configuration
    SLOT_MIN = 30   # minutes
    DAY_START_MIN = 8 * 60  # start at 08:00
    
    if not os.path.exists(temp_file):
        raise FileNotFoundError(f"فایل موقت یافت نشد: {temp_file}")
    
    print("در حال خواندن فایل موقت:", temp_file)
    xls = pd.ExcelFile(temp_file)
    print("شیت‌های یافت شده:", xls.sheet_names)
    
    # helper: normalize time string -> minutes
    def to_minutes(t):
        if pd.isna(t) or str(t).strip() == "":
            return None
        s = str(t).strip()
        s = s.translate(str.maketrans('۰۱۲۳۴۵۶۷۸۹','0123456789'))
        s = s.replace('.', ':').replace('：', ':')
        # if input like "8" -> "8:00"
        if ':' not in s and s.isdigit() and len(s) <= 2:
            try:
                return int(s) * 60
            except:
                return None
        if ':' in s:
            parts = s.split(':')
            try:
                h = int(parts[0])
                m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
                return h*60 + m
            except:
                return None
        # fallback try digits-only like "0830"
        if s.isdigit() and len(s) in (3,4):
            if len(s)==3: s = '0'+s
            hh = int(s[:-2]); mm = int(s[-2:])
            return hh*60 + mm
        return None
    
    def minute_label(m):
        hh = m//60; mm = m%60
        return f"{hh:02d}:{mm:02d}"
    
    # helper: find columns robustly
    def find_col(df_cols, candidates):
        for cand in candidates:
            for c in df_cols:
                if str(c).strip() == cand:
                    return c
        for cand in candidates:
            for c in df_cols:
                if cand in str(c):
                    return c
        return None
    
    # generate consistent light color based on course name
    def get_light_color(course_name):
        """Generate a consistent light pastel color based on course name"""
        if not course_name:
            return "FFFFFF"
        # Use hash to get consistent color for same course
        hash_val = int(hashlib.md5(course_name.encode()).hexdigest()[:8], 16)
        
        # Generate pastel colors using HSL technique (light colors)
        hues = [0, 30, 60, 120, 180, 240, 300]  # Red, Orange, Yellow, Green, Cyan, Blue, Magenta
        hue = hues[hash_val % len(hues)]
        
        # Light pastel colors (high lightness, medium saturation)
        if hue == 0:    # Red
            return "FFE6E6"  # Very light red
        elif hue == 30:  # Orange
            return "FFE8CC"  # Very light orange
        elif hue == 60:  # Yellow
            return "FFF9C4"  # Very light yellow
        elif hue == 120: # Green
            return "E6F7E6"  # Very light green
        elif hue == 180: # Cyan
            return "E6F7F7"  # Very light cyan
        elif hue == 240: # Blue
            return "E6E6FF"  # Very light blue
        else:           # Magenta
            return "F7E6F7"  # Very light magenta
    
    # build slots globally as needed per sheet (end depends on data)
    def build_slots(min_start, max_end):
        # ensure start is DAY_START_MIN
        start = DAY_START_MIN
        # round end up to nearest slot
        end = ((max_end + SLOT_MIN - 1)//SLOT_MIN)*SLOT_MIN
        if end <= start:
            end = start + 10 * 60  # fallback to 10 hours
        return list(range(start, end, SLOT_MIN))
    
    # collect which sheets we will build tables for
    weekday_names = ['شنبه','یکشنبه','دوشنبه','سه‌شنبه','چهارشنبه','پنج‌شنبه','جمعه']
    
    # Load the existing workbook (don't create a new one)
    wb = load_workbook(temp_file)
    
    # remove prior phase2 sheets if they exist (start fresh)
    for s in wb.sheetnames[:]:
        if s.startswith("جدول کلاسی "):
            wb.remove(wb[s])
    
    # iterate through Phase1 weekday sheets
    for sheet in xls.sheet_names:
        if sheet not in weekday_names:
            continue
        print("در حال پردازش شیت:", sheet)
        df = pd.read_excel(xls, sheet_name=sheet)
        if df.empty:
            print(" -> شیت خالی است، رد شد.")
            continue
        
        # find relevant columns robustly
        cols = list(df.columns)
        col_room = find_col(cols, ['مکان','نام مكان','مكان'])
        col_course = find_col(cols, ['نام درس','نام کلاس درس','نام کلاس'])
        col_teacher = find_col(cols, ['نام استاد','نام كامل استاد','PR S_FNAME','نام كامل'])
        col_code = find_col(cols, ['کد ارائه درس','کد ارائه','کد درس'])
        col_unit_th = find_col(cols, ['واحد نظری','تعداد واحد نظري','تعداد واحد'])
        col_unit_pr = find_col(cols, ['واحد عملی','تعداد واحد عملي'])
        col_group = find_col(cols, ['گروه آموزشی','نام گروه آموزشي','گروه'])
        col_degree = find_col(cols, ['مقطع'])
        col_reg = find_col(cols, ['تعداد ثبت نامی','تعداد ثبت نامي','تعداد ثبت نام'])
        col_M = find_col(cols, ['ساعت شروع','ساعت شروع کلاس','M','BV'])
        col_N = find_col(cols, ['ساعت پایان','ساعت پایان کلاس','N','BW'])
        
        if col_room is None:
            print(" -> ستون 'مکان' یافت نشد، رد شد.")
            continue
        
        # normalize textual columns
        for c in [col_room, col_course, col_teacher, col_code, col_unit_th, col_unit_pr, col_group, col_degree, col_reg]:
            if c is not None and c in df.columns:
                df[c] = df[c].fillna("").astype(str).str.replace('\u200c','').str.strip()
        # times
        if col_M in df.columns:
            df['_M_min'] = df[col_M].apply(to_minutes)
        else:
            df['_M_min'] = None
        if col_N in df.columns:
            df['_N_min'] = df[col_N].apply(to_minutes)
        else:
            df['_N_min'] = None
        
        # drop exact duplicates (same code, same room, same times)
        keycols = [c for c in [col_code, col_course, col_teacher, col_room, col_M, col_N] if c is not None]
        if keycols:
            df = df.drop_duplicates(subset=keycols)
        
        # determine slots (start at 08:00, end by max end)
        starts = df['_M_min'].dropna().tolist()
        ends = df['_N_min'].dropna().tolist()
        max_end = max(ends) if ends else (20*60)
        slots = build_slots(DAY_START_MIN, max_end)
        slot_labels = [minute_label(s) for s in slots]
        
        # prepare rooms: one row per unique room (exact string)
        rooms = df[col_room].fillna("").astype(str).unique().tolist()
        
        # NEW: Sort rooms by extracting numeric part
        def extract_number(room_name):
            """Extract numeric part from room name for sorting"""
            # Find all numbers in the string
            numbers = re.findall(r'\d+', room_name)
            if numbers:
                # Use the first number found
                return int(numbers[0])
            return 0  # Default for rooms without numbers
        
        # Sort rooms based on extracted number
        rooms.sort(key=lambda x: extract_number(x))
        
        # build a grid: dict room -> list per slot (None or list of entries)
        grid = {room: [None]*len(slots) for room in rooms}
        
        # fill grid: for each record mark slot indices that fully fit inside [M,N)
        for idx, row in df.iterrows():
            room = str(row[col_room])
            start = row.get('_M_min', None)
            end = row.get('_N_min', None)
            if start is None or end is None:
                continue
            
            # find start_idx: first slot s.t. slots[i] <= start < slots[i]+SLOT_MIN
            start_idx = None
            for i, s in enumerate(slots):
                if s <= start < s + SLOT_MIN:
                    start_idx = i
                    break
            if start_idx is None:
                start_idx = min(range(len(slots)), key=lambda k: abs(slots[k]-start))
            
            # end_idx: last index where slot_start + SLOT_MIN <= end (fully contained)
            end_idx = None
            for i, s in enumerate(slots):
                if s + SLOT_MIN <= end:
                    end_idx = i
            if end_idx is None or end_idx < start_idx:
                continue
            
            # Create unique entry identifier to avoid duplicates
            entry_id = f"{row[col_course] if col_course else ''}|{row[col_teacher] if col_teacher else ''}|{row[col_code] if col_code else ''}"
            
            # Create entry data
            entry_data = {
                'course': row[col_course] if col_course else "",
                'teacher': row[col_teacher] if col_teacher else "",
                'code': row[col_code] if col_code else "",
                'unit_th': row[col_unit_th] if col_unit_th else "",
                'unit_pr': row[col_unit_pr] if col_unit_pr else "",
                'group': row[col_group] if col_group else "",
                'degree': row[col_degree] if col_degree else "",
                'reg': row[col_reg] if col_reg else "",
                'M': row[col_M] if col_M else "",
                'N': row[col_N] if col_N else "",
                'entry_id': entry_id
            }
            
            # assign entry to each slot in range
            for k in range(start_idx, end_idx+1):
                if grid[room][k] is None:
                    grid[room][k] = []
                
                # Check if this exact entry already exists to avoid duplicates
                existing_entry_ids = [e['entry_id'] for e in grid[room][k]]
                if entry_id not in existing_entry_ids:
                    grid[room][k].append(entry_data)
        
        # Create phase2 sheet
        out_name = f"جدول کلاسی {sheet}"
        out_name = out_name[:31]
        ws = wb.create_sheet(title=out_name)
        
        # Title row merged
        total_cols = 1 + len(slot_labels)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"جدول کلاسی {sheet}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # header row (slot labels) in row 2
        ws.cell(row=2, column=1, value="مکان / ساعت").font = Font(bold=True)
        for j, lbl in enumerate(slot_labels, start=2):
            c = ws.cell(row=2, column=j, value=lbl)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9)
        
        # write room rows beginning at row 3
        start_row = 3
        for i, room in enumerate(rooms):
            r = start_row + i
            ws.cell(row=r, column=1, value=room)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 22
            
            # merge contiguous slots with same content
            j = 0
            while j < len(slots):
                cell_entries = grid[room][j]
                if not cell_entries:
                    j += 1
                    continue
                
                # Find contiguous slots with identical content
                k = j
                while k+1 < len(slots) and grid[room][k+1] == cell_entries:
                    k += 1
                
                excel_start = 2 + j
                excel_end = 2 + k
                
                # Merge cells
                if excel_end > excel_start:
                    ws.merge_cells(start_row=r, start_column=excel_start, end_row=r, end_column=excel_end)
                
                anchor = ws.cell(row=r, column=excel_start)
                
                # Display content (avoid duplicates)
                unique_entries = []
                seen_entry_ids = set()
                for ent in cell_entries:
                    if ent['entry_id'] not in seen_entry_ids:
                        unique_entries.append(ent)
                        seen_entry_ids.add(ent['entry_id'])
                
                # Format display text - only show unique entries
                display_lines = []
                tooltip_lines = []
                
                for ent in unique_entries:
                    display_line = f"{ent['course']} — {ent['teacher']}"
                    display_lines.append(display_line)
                    
                    # Simplified tooltip - removed گروه and مقطع to save space
                    tooltip_text = (
                        f"درس: {ent['course']}\n"
                        f"استاد: {ent['teacher']}\n"
                        f"کد: {ent['code']}\n"
                        f"واحد: {ent['unit_th']}(ن) + {ent['unit_pr']}(ع)\n"
                        f"ثبت‌نام: {ent['reg']}\n"
                        f"ساعت: {ent['M']} - {ent['N']}"
                    )
                    tooltip_lines.append(tooltip_text)
                
                # Only show unique display lines (avoid duplicates in display)
                unique_display_lines = list(set(display_lines))
                anchor.value = "\n".join(unique_display_lines)
                anchor.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                
                # Add tooltip comment with increased height
                if tooltip_lines:
                    try:
                        comment_text = "\n" + "─" * 30 + "\n".join(tooltip_lines)
                        anchor.comment = Comment(comment_text, "برنامه‌ساز")
                        anchor.comment.width = 350  # Increased width
                        anchor.comment.height = 200  # Increased height for better visibility
                    except Exception as e:
                        print(f"خطا در افزودن کامنت: {e}")
                
                # Apply light color based on course name
                if unique_entries:
                    first_course = unique_entries[0]['course']
                    color_hex = get_light_color(first_course)
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    anchor.fill = fill
                    
                    # Apply same fill to all merged cells
                    for col in range(excel_start, excel_end + 1):
                        ws.cell(row=r, column=col).fill = fill
                
                j = k + 1
        
        # Adjust column widths (reduced as requested)
        ws.column_dimensions[get_column_letter(1)].width = 25  # Reduced room column width
        for col_idx in range(2, 2 + len(slot_labels)):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 8  # Reduced from 20 to 8 (less than half)
        
        # center alignment for header area
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1+len(slot_labels)):
            for c in row:
                c.alignment = Alignment(horizontal="center", vertical="center")
    
    print("در حال ذخیره فایل نهایی:", final_output_file)
    wb.save(final_output_file)
    print("✅ انجام شد.")

def main():
    """Main function to run the complete process"""
    print("🎓 برنامه تولید جدول کلاسی")
    print("=" * 50)
    
    # Show welcome message first
    show_welcome_message()
    
    # Select input CSV file
    input_file = select_input_file()
    if not input_file:
        print("❌ هیچ فایلی انتخاب نشد.")
        return
    
    print(f"📁 فایل ورودی: {input_file}")
    
    # Select output Excel file
    output_file = select_output_file()
    if not output_file:
        print("❌ محل ذخیره فایل انتخاب نشد.")
        return
    
    print(f"📁 فایل خروجی: {output_file}")
    
    # Create temporary file in system temp directory to avoid access issues
    temp_file = os.path.join(tempfile.gettempdir(), "temp_schedule_phase1.xlsx")
    
    try:
        # Phase 1: Extract data from CSV
        print("\n🔹 مرحله 1: استخراج داده‌ها از فایل CSV...")
        if not phase1_extract_data(input_file, temp_file):
            return
        
        # Phase 2: Create schedule tables
        print("\n🔹 مرحله 2: ایجاد جداول کلاسی...")
        phase2_create_schedule(temp_file, output_file)
        
        print("\n🎉 برنامه با موفقیت به پایان رسید!")
        print(f"📊 فایل نهایی تولید شد: {output_file}")
        
        # Show success message
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("موفق", f"برنامه با موفقیت اجرا شد!\nفایل نهایی: {os.path.basename(output_file)}")
        
    except Exception as e:
        print(f"❌ خطا در اجرای برنامه: {e}")
        
        # Show error message
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("خطا", f"خطا در اجرای برنامه:\n{str(e)}")
        
    finally:
        # Clean up temporary file if it exists
        if os.path.exists(temp_file):
            try:
                # Make sure the file is closed before deleting
                import gc
                gc.collect()
                os.remove(temp_file)
                print(f"✅ فایل موقت پاک شد: {temp_file}")
            except Exception as e:
                print(f"⚠️ نتوانست فایل موقت را پاک کند: {e}")

if __name__ == "__main__":
    main()import gradio as gr
import pandas as pd
import tempfile
import os
import io
import hashlib
from math import ceil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
import re
import atexit
import glob

def cleanup_temp_files():
    """Clean up any remaining temporary files"""
    temp_files = glob.glob("/tmp/*_final.xlsx") + glob.glob("/tmp/*_phase1.xlsx")
    for temp_file in temp_files:
        try:
            if os.path.exists(temp_file):
                os.unlink(temp_file)
                print(f"🧹 Cleaned up: {temp_file}")
        except Exception as e:
            print(f"⚠️ Could not clean up {temp_file}: {e}")

# Register cleanup function
atexit.register(cleanup_temp_files)

def extract_time_from_calendar(calendar_text):
    """Extract start time from تقويم كلاس درس column"""
    if pd.isna(calendar_text) or not calendar_text:
        return None
    
    calendar_str = str(calendar_text).strip()
    
    # Pattern to match time in format "HH:MM تا HH:MM"
    time_pattern = r'(\d{1,2}:\d{2})\s*تا\s*\d{1,2}:\d{2}'
    match = re.search(time_pattern, calendar_str)
    
    if match:
        return match.group(1)  # Return the start time
    
    return None

def phase1_extract_data(input_file, temp_output_file):
    """Phase 1: Extract important data from CSV and save to Excel"""
    try:
        print("🔹 Phase 1: Starting data extraction...")
        
        # Read the uploaded file
        if hasattr(input_file, 'name'):  # Gradio file object
            file_path = input_file.name
        else:
            file_path = input_file
            
        print(f"🔹 Reading file: {file_path}")
        
        # Determine file type and read
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path, encoding='utf-8-sig')
            print("✅ CSV file read successfully")
        else:
            df = pd.read_excel(file_path)
            print("✅ Excel file read successfully")
        
        print(f"✅ File read successfully. Rows: {len(df)}, Columns: {len(df.columns)}")
        
        # ==== انتخاب ستون‌ها بر اساس شماره ====
        selected_columns = {
            'نام درس': 2,           # C
            'کد ارائه درس': 0,      # A
            'واحد نظری': 11,        # L
            'واحد عملی': 12,        # M
            'مکان': 22,             # W
            'گروه آموزشی': 43,      # AR
            'مقطع': 53,             # BB
            'تعداد ثبت نامی': 57,   # BF
            'نیم‌سال': 59,          # BH
            'نام استاد': 68,        # BQ
            'رشته': 70,             # BS
            'روز': 72,              # BU
            'ساعت شروع': 73,        # BV
            'ساعت پایان': 74,       # BW
            'تقويم كلاس درس': 71   # BT - برای استخراج زمان‌های خالی
        }
        
        # ==== استخراج فقط ستون‌های مورد نیاز ====
        df_selected = df.iloc[:, list(selected_columns.values())].copy()
        df_selected.columns = list(selected_columns.keys())
        
        # ==== پر کردن ساعت‌های شروع خالی از ستون تقويم كلاس درس ====
        print("🔹 Checking for empty start times...")
        empty_start_count = df_selected['ساعت شروع'].isna().sum()
        empty_start_count += (df_selected['ساعت شروع'] == '').sum()
        print(f"🔹 Found {empty_start_count} empty start times")
        
        if empty_start_count > 0:
            print("🔹 Filling empty start times from تقويم كلاس درس column...")
            filled_count = 0
            
            for idx, row in df_selected.iterrows():
                start_time = str(row['ساعت شروع']).strip() if pd.notna(row['ساعت شروع']) else ""
                calendar_text = row['تقويم كلاس درس']
                
                # If start time is empty but we have calendar data
                if not start_time and pd.notna(calendar_text) and str(calendar_text).strip():
                    extracted_time = extract_time_from_calendar(calendar_text)
                    if extracted_time:
                        df_selected.at[idx, 'ساعت شروع'] = extracted_time
                        filled_count += 1
                        print(f"   ↳ Filled row {idx}: {extracted_time}")
            
            print(f"✅ Filled {filled_count} empty start times from calendar data")
        
        # ==== پاکسازی و نرمال‌سازی ====
        def normalize_text(s):
            return (
                str(s)
                .replace('\u200c', '')   # حذف نیم‌فاصله
                .replace('ي', 'ی')       # ی عربی → فارسی
                .replace('ك', 'ک')       # ک عربی → فارسی
                .replace('‌', '')        # حذف ZWNJ اضافی
                .strip()
            )
        
        df_selected = df_selected.fillna("").astype(str)
        df_selected['روز'] = df_selected['روز'].apply(normalize_text)
        
        # ==== نگاشت دقیق اسامی روزها ====
        day_map = {
            'شنبه': 'شنبه',
            'یکشنبه': 'یکشنبه',
            'يکشنبه': 'یکشنبه',
            'يكشنبه': 'یکشنبه',
            'یكشنبه': 'یکشنبه',
            'دوشنبه': 'دوشنبه',
            'سه شنبه': 'سه‌شنبه',
            'سه‌شنبه': 'سه‌شنبه',
            'چهارشنبه': 'چهارشنبه',
            'چهار شنبه': 'چهارشنبه',
            'پنجشنبه': 'پنج‌شنبه',
            'پنج شنبه': 'پنج‌شنبه',
            'پنج‌شنبه': 'پنج‌شنبه',
            'پنچشنبه': 'پنج‌شنبه',
            'پنچ شنبه': 'پنج‌شنبه',
            'جمعه': 'جمعه'
        }
        
        df_selected['روز'] = df_selected['روز'].apply(
            lambda x: day_map[x] if x in day_map else x
        )
        
        # ==== لیست روزهای معتبر ====
        days = ['شنبه', 'یکشنبه', 'دوشنبه', 'سه‌شنبه', 'چهارشنبه', 'پنج‌شنبه', 'جمعه']
        
        # ==== تقسیم داده‌ها به شیت‌های مجزا و مرتب‌سازی ====
        sheets = {}
        for day in days:
            subset = df_selected[df_selected['روز'] == day].copy()
            if not subset.empty:
                # مرتب‌سازی بر اساس ساعت شروع
                # تبدیل ساعت شروع به عدد برای مرتب‌سازی
                def time_to_sortable(time_str):
                    if not time_str or str(time_str).strip() == "":
                        return 0
                    try:
                        # تبدیل زمان به دقیقه از ابتدای روز
                        time_str = str(time_str).strip()
                        time_str = time_str.translate(str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789'))
                        if ':' in time_str:
                            parts = time_str.split(':')
                            hours = int(parts[0])
                            minutes = int(parts[1]) if len(parts) > 1 else 0
                            return hours * 60 + minutes
                        else:
                            # اگر فقط عدد باشد (مثلاً "8")
                            return int(time_str) * 60
                    except:
                        return 0
                
                subset['ساعت شروع مرتب'] = subset['ساعت شروع'].apply(time_to_sortable)
                subset = subset.sort_values(by='ساعت شروع مرتب', ascending=True).drop(columns=['ساعت شروع مرتب'])
                sheets[day] = subset
        
        # ==== داده‌های با روز نامشخص ====
        unknown = df_selected[~df_selected['روز'].isin(days)]
        if not unknown.empty:
            sheets['نامشخص'] = unknown
        
        # ==== ذخیره در فایل اکسل ====
        with pd.ExcelWriter(temp_output_file, engine='openpyxl') as writer:
            for day, subset in sheets.items():
                # حذف ستون تقويم كلاس درس از خروجی نهایی
                subset_to_save = subset.drop(columns=['تقويم كلاس درس'], errors='ignore')
                subset_to_save.to_excel(writer, sheet_name=day[:30], index=False)
        
        print("✅ فایل اکسل موقت ساخته شد")
        return True
        
    except Exception as e:
        print(f"❌ خطا در فاز اول: {e}")
        import traceback
        print(f"🔍 Traceback:\n{traceback.format_exc()}")
        return False

def phase2_create_schedule(temp_file, final_output_file):
    """Phase 2: Create class schedule tables from the temporary Excel file"""
    
    # Configuration
    SLOT_MIN = 30   # minutes
    DAY_START_MIN = 8 * 60  # start at 08:00
    
    if not os.path.exists(temp_file):
        raise FileNotFoundError(f"فایل موقت یافت نشد: {temp_file}")
    
    print("در حال خواندن فایل موقت")
    xls = pd.ExcelFile(temp_file)
    print("شیت‌های یافت شده:", xls.sheet_names)
    
    # helper: normalize time string -> minutes
    def to_minutes(t):
        if pd.isna(t) or str(t).strip() == "":
            return None
        s = str(t).strip()
        s = s.translate(str.maketrans('۰۱۲۳۴۵۶۷۸۹','0123456789'))
        s = s.replace('.', ':').replace('：', ':')
        # if input like "8" -> "8:00"
        if ':' not in s and s.isdigit() and len(s) <= 2:
            try:
                return int(s) * 60
            except:
                return None
        if ':' in s:
            parts = s.split(':')
            try:
                h = int(parts[0])
                m = int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
                return h*60 + m
            except:
                return None
        # fallback try digits-only like "0830"
        if s.isdigit() and len(s) in (3,4):
            if len(s)==3: s = '0'+s
            hh = int(s[:-2]); mm = int(s[-2:])
            return hh*60 + mm
        return None
    
    def minute_label(m):
        hh = m//60; mm = m%60
        return f"{hh:02d}:{mm:02d}"
    
    # helper: find columns robustly
    def find_col(df_cols, candidates):
        for cand in candidates:
            for c in df_cols:
                if str(c).strip() == cand:
                    return c
        for cand in candidates:
            for c in df_cols:
                if cand in str(c):
                    return c
        return None
    
    # generate consistent light color based on course name
    def get_light_color(course_name):
        """Generate a consistent light pastel color based on course name"""
        if not course_name:
            return "FFFFFF"
        # Use hash to get consistent color for same course
        hash_val = int(hashlib.md5(course_name.encode()).hexdigest()[:8], 16)
        
        # Generate pastel colors using HSL technique (light colors)
        hues = [0, 30, 60, 120, 180, 240, 300]  # Red, Orange, Yellow, Green, Cyan, Blue, Magenta
        hue = hues[hash_val % len(hues)]
        
        # Light pastel colors (high lightness, medium saturation)
        if hue == 0:    # Red
            return "FFE6E6"  # Very light red
        elif hue == 30:  # Orange
            return "FFE8CC"  # Very light orange
        elif hue == 60:  # Yellow
            return "FFF9C4"  # Very light yellow
        elif hue == 120: # Green
            return "E6F7E6"  # Very light green
        elif hue == 180: # Cyan
            return "E6F7F7"  # Very light cyan
        elif hue == 240: # Blue
            return "E6E6FF"  # Very light blue
        else:           # Magenta
            return "F7E6F7"  # Very light magenta
    
    # build slots globally as needed per sheet (end depends on data)
    def build_slots(min_start, max_end):
        # ensure start is DAY_START_MIN
        start = DAY_START_MIN
        # round end up to nearest slot
        end = ((max_end + SLOT_MIN - 1)//SLOT_MIN)*SLOT_MIN
        if end <= start:
            end = start + 10 * 60  # fallback to 10 hours
        return list(range(start, end, SLOT_MIN))
    
    # collect which sheets we will build tables for
    weekday_names = ['شنبه','یکشنبه','دوشنبه','سه‌شنبه','چهارشنبه','پنج‌شنبه','جمعه']
    
    # Load the existing workbook (don't create a new one)
    wb = load_workbook(temp_file)
    
    # remove prior phase2 sheets if they exist (start fresh)
    for s in wb.sheetnames[:]:
        if s.startswith("جدول کلاسی "):
            wb.remove(wb[s])
    
    # iterate through Phase1 weekday sheets
    for sheet in xls.sheet_names:
        if sheet not in weekday_names:
            continue
        print("در حال پردازش شیت:", sheet)
        df = pd.read_excel(xls, sheet_name=sheet)
        if df.empty:
            print(" -> شیت خالی است، رد شد.")
            continue
        
        # find relevant columns robustly
        cols = list(df.columns)
        col_room = find_col(cols, ['مکان','نام مكان','مكان'])
        col_course = find_col(cols, ['نام درس','نام کلاس درس','نام کلاس'])
        col_teacher = find_col(cols, ['نام استاد','نام كامل استاد','PR S_FNAME','نام كامل'])
        col_code = find_col(cols, ['کد ارائه درس','کد ارائه','کد درس'])
        col_unit_th = find_col(cols, ['واحد نظری','تعداد واحد نظري','تعداد واحد'])
        col_unit_pr = find_col(cols, ['واحد عملی','تعداد واحد عملي'])
        col_group = find_col(cols, ['گروه آموزشی','نام گروه آموزشي','گروه'])
        col_degree = find_col(cols, ['مقطع'])
        col_reg = find_col(cols, ['تعداد ثبت نامی','تعداد ثبت نامي','تعداد ثبت نام'])
        col_M = find_col(cols, ['ساعت شروع','ساعت شروع کلاس','M','BV'])
        col_N = find_col(cols, ['ساعت پایان','ساعت پایان کلاس','N','BW'])
        
        if col_room is None:
            print(" -> ستون 'مکان' یافت نشد، رد شد.")
            continue
        
        # normalize textual columns
        for c in [col_room, col_course, col_teacher, col_code, col_unit_th, col_unit_pr, col_group, col_degree, col_reg]:
            if c is not None and c in df.columns:
                df[c] = df[c].fillna("").astype(str).str.replace('\u200c','').str.strip()
        # times
        if col_M in df.columns:
            df['_M_min'] = df[col_M].apply(to_minutes)
        else:
            df['_M_min'] = None
        if col_N in df.columns:
            df['_N_min'] = df[col_N].apply(to_minutes)
        else:
            df['_N_min'] = None
        
        # drop exact duplicates (same code, same room, same times)
        keycols = [c for c in [col_code, col_course, col_teacher, col_room, col_M, col_N] if c is not None]
        if keycols:
            df = df.drop_duplicates(subset=keycols)
        
        # determine slots (start at 08:00, end by max end)
        starts = df['_M_min'].dropna().tolist()
        ends = df['_N_min'].dropna().tolist()
        max_end = max(ends) if ends else (20*60)
        slots = build_slots(DAY_START_MIN, max_end)
        slot_labels = [minute_label(s) for s in slots]
        
        # prepare rooms: one row per unique room (exact string)
        rooms = df[col_room].fillna("").astype(str).unique().tolist()
        # build a grid: dict room -> list per slot (None or list of entries)
        grid = {room: [None]*len(slots) for room in rooms}
        
        # fill grid: for each record mark slot indices that fully fit inside [M,N)
        for idx, row in df.iterrows():
            room = str(row[col_room])
            start = row.get('_M_min', None)
            end = row.get('_N_min', None)
            if start is None or end is None:
                continue
            
            # find start_idx: first slot s.t. slots[i] <= start < slots[i]+SLOT_MIN
            start_idx = None
            for i, s in enumerate(slots):
                if s <= start < s + SLOT_MIN:
                    start_idx = i
                    break
            if start_idx is None:
                start_idx = min(range(len(slots)), key=lambda k: abs(slots[k]-start))
            
            # end_idx: last index where slot_start + SLOT_MIN <= end (fully contained)
            end_idx = None
            for i, s in enumerate(slots):
                if s + SLOT_MIN <= end:
                    end_idx = i
            if end_idx is None or end_idx < start_idx:
                continue
            
            # Create unique entry identifier to avoid duplicates
            entry_id = f"{row[col_course] if col_course else ''}|{row[col_teacher] if col_teacher else ''}|{row[col_code] if col_code else ''}"
            
            # Create entry data
            entry_data = {
                'course': row[col_course] if col_course else "",
                'teacher': row[col_teacher] if col_teacher else "",
                'code': row[col_code] if col_code else "",
                'unit_th': row[col_unit_th] if col_unit_th else "",
                'unit_pr': row[col_unit_pr] if col_unit_pr else "",
                'group': row[col_group] if col_group else "",
                'degree': row[col_degree] if col_degree else "",
                'reg': row[col_reg] if col_reg else "",
                'M': row[col_M] if col_M else "",
                'N': row[col_N] if col_N else "",
                'entry_id': entry_id
            }
            
            # assign entry to each slot in range
            for k in range(start_idx, end_idx+1):
                if grid[room][k] is None:
                    grid[room][k] = []
                
                # Check if this exact entry already exists to avoid duplicates
                existing_entry_ids = [e['entry_id'] for e in grid[room][k]]
                if entry_id not in existing_entry_ids:
                    grid[room][k].append(entry_data)
        
        # Create phase2 sheet
        out_name = f"جدول کلاسی {sheet}"
        out_name = out_name[:31]
        ws = wb.create_sheet(title=out_name)
        
        # Title row merged
        total_cols = 1 + len(slot_labels)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"جدول کلاسی {sheet}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # header row (slot labels) in row 2
        ws.cell(row=2, column=1, value="مکان / ساعت").font = Font(bold=True)
        for j, lbl in enumerate(slot_labels, start=2):
            c = ws.cell(row=2, column=j, value=lbl)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(size=9)
        
        # write room rows beginning at row 3
        start_row = 3
        for i, room in enumerate(rooms):
            r = start_row + i
            ws.cell(row=r, column=1, value=room)
            ws.cell(row=r, column=1).alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[r].height = 22
            
            # merge contiguous slots with same content
            j = 0
            while j < len(slots):
                cell_entries = grid[room][j]
                if not cell_entries:
                    j += 1
                    continue
                
                # Find contiguous slots with identical content
                k = j
                while k+1 < len(slots) and grid[room][k+1] == cell_entries:
                    k += 1
                
                excel_start = 2 + j
                excel_end = 2 + k
                
                # Merge cells
                if excel_end > excel_start:
                    ws.merge_cells(start_row=r, start_column=excel_start, end_row=r, end_column=excel_end)
                
                anchor = ws.cell(row=r, column=excel_start)
                
                # Display content (avoid duplicates)
                unique_entries = []
                seen_entry_ids = set()
                for ent in cell_entries:
                    if ent['entry_id'] not in seen_entry_ids:
                        unique_entries.append(ent)
                        seen_entry_ids.add(ent['entry_id'])
                
                # Format display text - only show unique entries
                display_lines = []
                tooltip_lines = []
                
                for ent in unique_entries:
                    display_line = f"{ent['course']} — {ent['teacher']}"
                    display_lines.append(display_line)
                    
                    # Simplified tooltip
                    tooltip_text = (
                        f"درس: {ent['course']}\n"
                        f"استاد: {ent['teacher']}\n"
                        f"کد: {ent['code']}\n"
                        f"واحد: {ent['unit_th']}(ن) + {ent['unit_pr']}(ع)\n"
                        f"ثبت‌نام: {ent['reg']}\n"
                        f"ساعت: {ent['M']} - {ent['N']}"
                    )
                    tooltip_lines.append(tooltip_text)
                
                # Only show unique display lines (avoid duplicates in display)
                unique_display_lines = list(set(display_lines))
                anchor.value = "\n".join(unique_display_lines)
                anchor.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
                
                # Add tooltip comment with increased height
                if tooltip_lines:
                    try:
                        comment_text = "\n" + "─" * 30 + "\n".join(tooltip_lines)
                        anchor.comment = Comment(comment_text, "برنامه‌ساز")
                        anchor.comment.width = 350
                        anchor.comment.height = 200
                    except Exception as e:
                        print(f"خطا در افزودن کامنت: {e}")
                
                # Apply light color based on course name
                if unique_entries:
                    first_course = unique_entries[0]['course']
                    color_hex = get_light_color(first_course)
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    anchor.fill = fill
                    
                    # Apply same fill to all merged cells
                    for col in range(excel_start, excel_end + 1):
                        ws.cell(row=r, column=col).fill = fill
                
                j = k + 1
        
        # Adjust column widths
        ws.column_dimensions[get_column_letter(1)].width = 25
        for col_idx in range(2, 2 + len(slot_labels)):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 8
        
        # center alignment for header area
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=1+len(slot_labels)):
            for c in row:
                c.alignment = Alignment(horizontal="center", vertical="center")
    
    print("در حال ذخیره فایل نهایی")
    wb.save(final_output_file)
    print("✅ انجام شد.")

def process_file(file):
    """Process the uploaded file and return download link"""
    temp_phase1 = None
    temp_final = None
    
    try:
        print("🔹 Starting file processing...")
        
        # Create temporary files
        with tempfile.NamedTemporaryFile(delete=False, suffix='_phase1.xlsx') as tmp1:
            temp_phase1 = tmp1.name
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='_final.xlsx') as tmp2:
            temp_final = tmp2.name
        
        print(f"🔹 Temporary files created: {temp_phase1}, {temp_final}")
        
        # Run phase 1
        print("🔹 Starting Phase 1...")
        if phase1_extract_data(file, temp_phase1):
            print("✅ Phase 1 completed successfully")
            
            # Run phase 2
            print("🔹 Starting Phase 2...")
            phase2_create_schedule(temp_phase1, temp_final)
            print("✅ Phase 2 completed successfully")
            
            # Return the file path, not the bytes data
            print(f"✅ Processing complete. Final file: {temp_final}")
            return temp_final, "جدول_کلاسی_نهایی.xlsx"
        else:
            print("❌ Phase 1 failed")
            return None, "خطا در پردازش فاز اول"
            
    except Exception as e:
        print(f"❌ Error in process_file: {str(e)}")
        import traceback
        error_details = traceback.format_exc()
        print(f"🔍 Full traceback:\n{error_details}")
        return None, f"خطا: {str(e)}"
    
    finally:
        # Clean up temporary files (except the final one which Gradio will handle)
        if temp_phase1 and os.path.exists(temp_phase1):
            try:
                os.unlink(temp_phase1)
                print("✅ Phase 1 temp file cleaned up")
            except Exception as e:
                print(f"⚠️ Could not delete phase1 temp file: {e}")

# Create the interface with Persian RTL layout
with gr.Blocks(
    title="برنامه جدول کلاسی",
    theme=gr.themes.Soft(),
    css="""
    .container {
        direction: rtl;
        text-align: right;
        font-family: Tahoma;
    }
    """
) as demo:
    
    gr.Markdown("""
    # 🎓 برنامه تولید جدول کلاسی دانشگاه
    **نسخه 1 - آبان 1404 - نیماوزیری**
    
    لطفا فایل خروجی آموزشیار (CSV) را آپلود کنید
    """)
    
    with gr.Row():
        with gr.Column(scale=1):
            file_input = gr.File(
                label="📁 آپلود فایل",
                file_types=[".csv", ".xlsx"],
                type="filepath"
            )
            
            process_btn = gr.Button(
                "🚀 شروع پردازش",
                variant="primary",
                size="lg"
            )
    
    with gr.Row():
        with gr.Column(scale=1):
            status_display = gr.Textbox(
                label="وضعیت",
                interactive=False,
                value="در انتظار آپلود فایل...",
                lines=2
            )
            
            download_output = gr.File(
                label="📥 دانلود فایل خروجی",
                file_types=[".xlsx"],
                visible=False
            )
    
    # Process function
    def process_and_update(file):
        if file is None:
            return "لطفا ابتدا فایل را آپلود کنید", None
        
        try:
            file_path, filename = process_file(file)
            if file_path and os.path.exists(file_path):
                return "✅ پردازش با موفقیت انجام شد!", gr.update(value=file_path, label=filename, visible=True)
            else:
                return f"❌ {filename}", gr.update(visible=False)
                
        except Exception as e:
            error_msg = f"❌ خطا: {str(e)}"
            print(f"Final error: {error_msg}")
            return error_msg, gr.update(visible=False)
    
    process_btn.click(
        fn=process_and_update,
        inputs=file_input,
        outputs=[status_display, download_output]
    )
    
    # Add a cleanup trigger when the download is used
    def cleanup_after_download():
        """Clean up files after some time"""
        import time
        time.sleep(60)  # Wait 60 seconds before cleanup
        cleanup_temp_files()
    
    # You can trigger cleanup when new file is uploaded
    file_input.change(
        fn=lambda: cleanup_temp_files(),
        inputs=None,
        outputs=None
    )

if __name__ == "__main__":
    demo.launch()